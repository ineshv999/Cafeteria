from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.usuario import Usuario
from app.models.producto import Producto
from app.models.categoria import Categoria
from app.models.pedido import Pedido
from app.models.detalle_pedido import DetallePedido

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from openpyxl import Workbook

from fastapi.responses import FileResponse

from sqlalchemy.orm import joinedload

import tempfile

class ReporteService:

    @staticmethod
    def obtener(db: Session):

        total_usuarios = db.query(Usuario).count()

        total_productos = db.query(Producto).count()

        total_categorias = db.query(Categoria).count()

        total_pedidos = db.query(Pedido).count()

        ventas = (
            db.query(func.sum(Pedido.total))
            .scalar()
            or 0
        )

        poco_stock = (
            db.query(Producto)
            .filter(Producto.stock <= 5)
            .count()
        )

        productos_vendidos = (
            db.query(
                func.sum(DetallePedido.cantidad)
            ).scalar()
            or 0
        )

        return {

            "usuarios": total_usuarios,

            "productos": total_productos,

            "categorias": total_categorias,

            "pedidos": total_pedidos,

            "ventas": float(ventas),

            "poco_stock": poco_stock,

            "productos_vendidos": int(productos_vendidos)

        }

    @staticmethod
    def generar_pdf(db: Session):

        datos = ReporteService.obtener(db)

        archivo = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".pdf"
        )

        pdf = SimpleDocTemplate(archivo.name)

        tabla = Table([

            ["Concepto","Valor"],

            ["Usuarios", datos["usuarios"]],

            ["Productos", datos["productos"]],

            ["Categorias", datos["categorias"]],

            ["Pedidos", datos["pedidos"]],

            ["Ventas", f"${datos['ventas']:.2f}"],

            ["Productos vendidos", datos["productos_vendidos"]],

            ["Stock bajo", datos["poco_stock"]]

        ])

        tabla.setStyle(TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.brown),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("BOTTOMPADDING",(0,0),(-1,0),10)

        ]))

        pdf.build([tabla])

        return FileResponse(

            archivo.name,

            filename="reporte.pdf",

            media_type="application/pdf"

        )

    @staticmethod
    def generar_excel(db: Session):

        datos = ReporteService.obtener(db)

        wb = Workbook()

        ws = wb.active

        ws.title = "Reporte"

        ws.append(["Concepto","Valor"])

        ws.append(["Usuarios",datos["usuarios"]])

        ws.append(["Productos",datos["productos"]])

        ws.append(["Categorias",datos["categorias"]])

        ws.append(["Pedidos",datos["pedidos"]])

        ws.append(["Ventas",datos["ventas"]])

        ws.append(["Productos vendidos",datos["productos_vendidos"]])

        ws.append(["Stock bajo",datos["poco_stock"]])

        archivo = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        )

        wb.save(archivo.name)

        return FileResponse(

            archivo.name,

            filename="reporte.xlsx",

            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        )
    
    @staticmethod
    def reporte_productos(

        db: Session,

        nombre=None,

        id_categoria=None,

        activo=None,

        stock_minimo=None

    ):

        consulta = (
            db.query(Producto)
            .options(joinedload(Producto.categoria))
        )

        if nombre:

            consulta = consulta.filter(
                Producto.nombre.ilike(f"%{nombre}%")
            )

        if id_categoria:

            consulta = consulta.filter(
                Producto.id_categoria == id_categoria
            )

        if activo is not None:

            consulta = consulta.filter(
                Producto.activo == activo
            )

        if stock_minimo:

            consulta = consulta.filter(
                Producto.stock <= stock_minimo
            )

        productos = consulta.all()

        for p in productos:
            print("================================")
            print("Producto:", p.nombre)
            print("Categoria:", p.categoria)
            print("Nombre categoria:", p.categoria.nombre)

        return productos
    
    @staticmethod
    def generar_pdf_productos(

        db: Session,

        nombre=None,

        categoria=None,

        activo=None,

        stock_minimo=None

    ):

        productos = ReporteService.listar_productos(

            db,

            nombre,

            categoria,

            activo,

            stock_minimo

        )

        archivo = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".pdf"
        )

        pdf = SimpleDocTemplate(archivo.name)

        datos = [[

            "Producto",

            "Categoría",

            "Precio",

            "Stock",

            "Estado"

        ]]

        for p in productos:

            datos.append([

                p.nombre,

                p.categoria.nombre,

                f"${float(p.precio):.2f}",

                p.stock,

                "Activo" if p.activo else "Inactivo"

            ])

        tabla = Table(datos)

        tabla.setStyle(TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.brown),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")

        ]))

        pdf.build([tabla])

        return FileResponse(

            archivo.name,

            filename="productos.pdf",

            media_type="application/pdf"

        )
    
    @staticmethod
    def generar_excel_productos(

        db: Session,

        nombre=None,

        categoria=None,

        activo=None,

        stock_minimo=None

    ):

        productos = ReporteService.reporte_productos(

            db,

            nombre,

            categoria,

            activo,

            stock_minimo

        )

        wb = Workbook()

        ws = wb.active

        ws.title = "Productos"

        ws.append([

            "Producto",

            "Categoría",

            "Precio",

            "Stock",

            "Estado"

        ])

        for p in productos:

            ws.append([

                p.nombre,

                p.categoria.nombre,

                f"${float(p.precio):.2f}",

                p.stock,

                "Activo" if p.activo else "Inactivo"

            ])

        archivo = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        )

        wb.save(archivo.name)

        return FileResponse(

            archivo.name,

            filename="productos.xlsx",

            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        )
    
    @staticmethod
    def reporte_pedidos(

        db: Session,

        estado=None,

        id_mesa=None,

        fecha_inicio=None,

        fecha_fin=None

    ):

        consulta = (
            db.query(Pedido)
            .join(Pedido.mesa)
            .join(Pedido.usuario)
        )

        if estado:

            consulta = consulta.filter(
                Pedido.estado == estado
            )

        if id_mesa:

            consulta = consulta.filter(
                Pedido.id_mesa == id_mesa
            )

        if fecha_inicio:

            consulta = consulta.filter(
                Pedido.fecha >= fecha_inicio
            )

        if fecha_fin:

            consulta = consulta.filter(
                Pedido.fecha <= fecha_fin
            )

        return consulta.order_by(
            Pedido.fecha.desc()
        ).all()
    
    @staticmethod
    def reporte_inventario(

        db: Session,

        categoria=None,

        stock_bajo=None

    ):

        consulta = db.query(Producto)

        if categoria:

            consulta = consulta.filter(
                Producto.id_categoria == categoria
            )

        if stock_bajo:

            consulta = consulta.filter(
                Producto.stock <= 5
            )

        return consulta.all()

    @staticmethod
    def listar_productos(
        db: Session,
        nombre=None,
        categoria=None,
        activo=None,
        stock_minimo=None
    ):

        consulta = (
            db.query(Producto)
            .options(joinedload(Producto.categoria))
        )

        if nombre:

            consulta = consulta.filter(
                Producto.nombre.ilike(f"%{nombre}%")
            )

        if categoria:

            consulta = consulta.filter(
                Producto.id_categoria == categoria
            )

        if activo is not None:

            if isinstance(activo, str):
                activo = activo.lower() == "true"

            consulta = consulta.filter(
                Producto.activo == activo
            )

        if stock_minimo:

            consulta = consulta.filter(
                Producto.stock <= int(stock_minimo)
            )

        return consulta.order_by(
            Producto.nombre
        ).all()
    
    @staticmethod
    def listar_pedidos(

        db: Session,

        estado=None,

        id_mesa=None,

        fecha_inicio=None,

        fecha_fin=None

    ):

        consulta = (

            db.query(Pedido)

            .options(

                joinedload(Pedido.mesa),

                joinedload(Pedido.usuario)

            )

        )

        if estado:

            consulta = consulta.filter(
                Pedido.estado == estado
            )

        if id_mesa:

            consulta = consulta.filter(
                Pedido.id_mesa == id_mesa
            )

        if fecha_inicio:

            consulta = consulta.filter(
                Pedido.fecha >= fecha_inicio
            )

        if fecha_fin:

            consulta = consulta.filter(
                Pedido.fecha <= fecha_fin
            )

        return consulta.order_by(
            Pedido.fecha.desc()
        ).all()
    
    @staticmethod
    def generar_pdf_pedidos(

        db: Session,

        estado=None,

        id_mesa=None,

        fecha_inicio=None,

        fecha_fin=None

    ):

        pedidos = ReporteService.listar_pedidos(

            db,

            estado,

            id_mesa,

            fecha_inicio,

            fecha_fin

        )

        archivo = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".pdf"
        )

        pdf = SimpleDocTemplate(archivo.name)

        datos = [[

            "Pedido",

            "Mesa",

            "Estado",

            "Total",

            "Fecha"

        ]]

        for p in pedidos:

            datos.append([

                p.id_pedido,

                p.mesa.numero,

                p.estado,

                f"${float(p.total):.2f}",

                p.fecha.strftime("%d/%m/%Y")

            ])

        tabla = Table(datos)

        tabla.setStyle(TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.brown),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")

        ]))

        pdf.build([tabla])

        return FileResponse(

            archivo.name,

            filename="pedidos.pdf",

            media_type="application/pdf"

        )

    @staticmethod
    def generar_excel_pedidos(

        db: Session,

        estado=None,

        id_mesa=None,

        fecha_inicio=None,

        fecha_fin=None

    ):

        pedidos = ReporteService.listar_pedidos(

            db,

            estado,

            id_mesa,

            fecha_inicio,

            fecha_fin

        )

        wb = Workbook()

        ws = wb.active

        ws.title = "Pedidos"

        ws.append([

            "Pedido",

            "Mesa",

            "Estado",

            "Total",

            "Fecha"

        ])

        for p in pedidos:

            ws.append([

                p.id_pedido,

                p.mesa.numero,

                p.estado,

                float(p.total),

                p.fecha.strftime("%d/%m/%Y")

            ])

        archivo = tempfile.NamedTemporaryFile(

            delete=False,

            suffix=".xlsx"

        )

        wb.save(archivo.name)

        return FileResponse(

            archivo.name,

            filename="pedidos.xlsx",

            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        )